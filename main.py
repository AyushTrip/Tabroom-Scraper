#Import Nessecary Libraries
#PD for data-table scraping
import pandas as pd
import requests
import json
from bs4 import BeautifulSoup

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def find_speaks(x, speak):
  
  nums = []
  
  for i in range(len(x)):
    if(x[i].isdigit()):
      nums.append(int(x[i]))

  if (speak):
    first_speak = float(nums[0] * 10) + float(nums[1]) + float(nums[2]*0.1)
    second_speak = float(nums[4] * 10) + float(nums[5]) + float(nums[6]*0.1)
    
  else:
    first_speak = float(nums[0] * 10) + float(nums[1]) + float(nums[2]*0.1)
    second_speak = float(nums[3] * 10) + float(nums[4]) + float(nums[5]*0.1)

  return (first_speak + second_speak) / 2

#Take URL and Round Input
url = str(input("Enter Round 1 Tabroom URL: "))
rounds = int(input("Enter # of Prelims: "))

#Find out tournament result style
ask_speak = str(input("Does The Tournament List Speaker Positions in the Results Category y/n: "))
speak_pos = False
if(ask_speak == 'y'): speak_pos = True

#Generate all round links
last_part_id = int(url[-2:])
round_links = []
temp_rounds = 0

for i in range(rounds):
  round_links.append(url[0:-2]+str(last_part_id+temp_rounds))
  temp_rounds += 1

#Create Dictionaries
final_speaks = dict()
rounds_judged = dict()

#Open Links
#SCRAPING PROCESS ============================================
try:
  
  for i in range(len(round_links)):

    round_html = requests.get(round_links[i])
    round_soup = BeautifulSoup(round_html.text, 'html.parser')
    dfs = pd.read_html(round_html.text)
    table = dfs[0]

    #Iterate through rows
    for j in range(len(table)):

      try:
        
        t1 = find_speaks(str(table.iloc[j,4]), speak_pos)
        t2 = find_speaks(str(table.iloc[j,5]), speak_pos)
        judge_name = table.iloc[j,2]
        average = float(t1+t2)/2

        #Check if judge is in the dictionary
        
        #If NO
        if judge_name not in final_speaks:
          
          final_speaks.update({judge_name:average})
          rounds_judged.update({judge_name:1})
          
        #If YES
        else:
          
          current_avg = final_speaks.get(judge_name)
          current_rounds = rounds_judged.get(judge_name) 
          final_speaks[judge_name] = current_avg + average
          rounds_judged[judge_name] = current_rounds + 1
          
      except:
        
        pass

except:
  
  pass

#Calculate Average
for key, value in final_speaks.items():
  rounds = rounds_judged.get(key)
  final_speaks[key] = round(float(value/rounds),4)

#SPREADSHEET TRANSFER ===========================
wb = load_workbook('tabroom-automation/yao.xlsx')
ws = wb.active

#Iterate through all the judges in the dictionary
for key, value in final_speaks.items():

    value_done = False
    row = 2
    col = 1

    while not value_done:

      if ws.cell(row=row,column=col).value == str(key):

          iteration = 1 
          found_empty = False

          while found_empty == False:
              
              if ws.cell(row=row, column=col+iteration).value is None:
                  ws.cell(row=row, column=col+iteration).value = value
                  found_empty = True 
                  value_done = True

              iteration += 1
              
      elif ws.cell(row=row,column=col).value is None:
        
          ws.cell(row=row,column=col).value = str(key)
          ws.cell(row=row,column=col+1).value = value
          value_done = True
          
      else:
        
        row += 1

wb.save('tabroom-automation/yao.xlsx')

#PROPERTY OF WESTWOOD DEBATE
